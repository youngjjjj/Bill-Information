import openpyxl
from bs4 import BeautifulSoup
from urllib.parse import unquote
import requests
import json

with open("secret.json") as f:
    secret = json.loads(f.read())

# Keep secret keys in secret.json
def get_secret(setting, secret=secret):
    try:
        return secret[setting]
    except KeyError:
        error_msg = "Set the {0} environment variable".format(setting)
        raise Exception(error_msg)


excel_file = openpyxl.load_workbook(
    "국회의원 입법활동 조사 명단(대표발의).xlsx"
)

sheet_person = excel_file["명단"]

sheet_result = excel_file["결과"]

API_KEY = get_secret("API_KEY")

url = "http://apis.data.go.kr/9710000/BillInfoService2/getBillInfoList"

decode_API_KEY = unquote(API_KEY)

people_list = [row[1].value for row in sheet_person.rows]

idx = 1
count = 1

for person in people_list:
    params = {
        "ServiceKey": decode_API_KEY,
        "pageNo": "1",
        "numOfRows": "700",
        "mem_name_check": "G01",
        "mem_name": person,
        "start_ord": "20",
        "end_ord": "20",
        "process_num": "-",
        "start_process_num": "-",
        "end_process_num": "-",
        "propose_num": "-",
        "start_propose_num": "-",
        "end_propose_num": "-",
        "proposer_kind_cd": "F01",
        "gbn": "dae_num_name",
    }

    response = requests.get(url, params=params)
    soup = BeautifulSoup(response.text, "html.parser")
    content = soup.find("response")
    items = content.find("items")

    sheet_person.cell(row=count, column=3).value = len(items)
    count += 1

    for item in items:
        # 의원명
        sheet_result.cell(row=idx + 1, column=1).value = person
        # 의안명
        sheet_result.cell(row=idx + 1, column=2).value = (
            item.find("billname").text if not item.find("billname") is None else ""
        )
        # 제안일자
        sheet_result.cell(row=idx + 1, column=3).value = (
            item.find("proposedt").text if not item.find("proposedt") is None else ""
        )
        # 의결결과
        sheet_result.cell(row=idx + 1, column=4).value = (
            item.find("generalresult").text
            if not item.find("generalresult") is None
            else ""
        )
        # 주요내용
        sheet_result.cell(row=idx + 1, column=5).value = (
            item.find("summary").text if not item.find("summary") is None else ""
        )
        # 심사진행 상태
        sheet_result.cell(row=idx + 1, column=6).value = (
            item.find("procstagecd").text
            if not item.find("procstagecd") is None
            else ""
        )
        idx += 1
    idx += 1
    print(f"{len(people_list)}명 중 {people_list.index(person) + 1}번째 입니다.")

excel_file.save(filename="국회의원 입법활동 조사 명단(대표발의).xlsx")
print("Complete!")
